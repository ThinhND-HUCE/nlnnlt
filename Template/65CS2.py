from openpyxl import load_workbook, Workbook
import random as rd

# Đọc
wb = load_workbook('Programming_Language_List.xlsx')
sh = wb.active

languages = []
for i in range(1, sh.max_row + 1):
    v = sh.cell(row=i, column=1).value
    if type(v) == int or type(v) == str and v.isnumeric():
        languages.append(sh.cell(row=i, column=2).value)

permutation_of_languages = []
while True:
    n = len(languages)
    i = rd.randint(0, n-1)
    x = languages.pop(i)
    permutation_of_languages.append(x)
    if not languages:
        break

print(permutation_of_languages)
wb.close()

# Ghi
wb = Workbook()
sh = wb.active
sh.cell(row=1, column=1, value='Nhóm')
sh.cell(row=1, column=2, value='Ngôn ngữ')

for i in range(len(permutation_of_languages)):
    sh.cell(row=i+2, column=1, value=i+1)
    sh.cell(row=i+2, column=2, value=permutation_of_languages[i])

wb.save('Assignment_List.xlsx')
wb.close()
